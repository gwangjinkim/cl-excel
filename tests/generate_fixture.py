# /// script
# dependencies = ["openpyxl"]
# ///

import openpyxl
import datetime
import os

def create_fixture():
    # Ensure fixtures directory exists
    os.makedirs("tests/fixtures", exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header
    ws["A1"] = "String"
    ws["B1"] = "Integer"
    ws["C1"] = "Float"
    ws["D1"] = "Boolean"
    ws["E1"] = "Date"
    ws["F1"] = "DateTime"
    ws["G1"] = "Missing"

    # Row 2: Basic values
    ws["A2"] = "Hello"
    ws["B2"] = 42
    ws["C2"] = 3.14
    ws["D2"] = True
    ws["E2"] = datetime.date(2023, 10, 27)
    ws["F2"] = datetime.datetime(2023, 10, 27, 12, 30, 45)
    ws["G2"] = None # Missing

    # Row 3: More values
    ws["A3"] = "World"
    ws["B3"] = -100
    ws["C3"] = -0.001
    ws["D3"] = False
    ws["E3"] = datetime.date(2024, 1, 1)
    ws["F3"] = datetime.datetime(2024, 1, 1, 0, 0, 0)
    # G3 is explicitly empty

    filename = "tests/fixtures/basic_types.xlsx"
    wb.save(filename)
    print(f"Generated {filename}")

if __name__ == "__main__":
    create_fixture()
